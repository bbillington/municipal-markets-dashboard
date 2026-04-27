"""
build_dashboard_data.py — Excel → per-city JSON pipeline (v2.27).

Reads each city's Contract Logs/<slug>_contracts.xlsx Config sheet, applies
Processing/firm_filters.classify_contract() as the single source of truth
for arch / non-engineering / MT / facility drops, and writes:

    docs/data/<slug>.json   ← per-city contracts (keep verdicts only)
    docs/data/manifest.json ← city directory + status summary
    docs/data/config.json   ← firm lists, category palette, halff aliases
                              (everything the dashboard used to hardcode)

Rationale for the per-city split (Item 27):
- One corrupted xlsx / locked file / schema drift fails ONE city, not the
  whole dashboard. Previously contract_data.json was all-or-nothing.
- Dashboard does Promise.allSettled on manifest.cities — rejected cities
  show as a muted pill in the footer, the other 10 render normally.
- Tiny files (~50 KB each) parallel-load in <100 ms on a typical connection.

Rationale for config.json:
- All firm classification rules (ARCHITECTURE_FIRMS, NON_ENGINEERING_FIRMS,
  palette, halff aliases) are now derived from a single source of truth:
  Processing/firm_filters.py + Processing/config.py. The dashboard stops
  carrying stale parallel lists.

Usage:
    py docs/build_dashboard_data.py
"""

from __future__ import annotations

import json
import os
import re
import sys
from collections import Counter, defaultdict
from datetime import datetime, date, timezone

# Ensure UTF-8 output on Windows
if sys.platform == "win32":
    sys.stdout.reconfigure(encoding="utf-8")

# ---------- paths ----------
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
REPO_ROOT = os.path.dirname(SCRIPT_DIR)
CONFIG_PATH = os.path.join(SCRIPT_DIR, "cities_config.json")

# Single source of truth: Processing/ modules for paths + filters
sys.path.insert(0, os.path.join(REPO_ROOT, "Processing"))
from config import CONTRACT_LOGS_DIR  # noqa: E402
from firm_filters import (  # noqa: E402
    ARCHITECTURE_FIRMS,
    NON_ENGINEERING_FIRMS,
    MATERIALS_GEOTECH_FIRMS,
    classify_contract,
)

OUTPUT_DIR = os.path.join(SCRIPT_DIR, "data")
MANIFEST_PATH = os.path.join(OUTPUT_DIR, "manifest.json")
DASHBOARD_CONFIG_PATH = os.path.join(OUTPUT_DIR, "config.json")
PMS_PATH = os.path.join(OUTPUT_DIR, "pms.json")

MIN_YEAR = 2023

# ── Project-type normalization (matches dashboard category palette below) ───
TYPE_MAP = {
    "drainage / floodplain / stormwater": "Drainage",
    "drainage/floodplain/stormwater": "Drainage",
    "floodplain": "Drainage",
    "stormwater": "Drainage",
    "flood mitigation": "Drainage",
    "hydraulic": "Drainage",
    "hydrology": "Drainage",
    "hydrologic": "Drainage",
    "h&h": "Drainage",
    "roads": "Roadway",
    "road": "Roadway",
    "roadway": "Roadway",
    "planning": "Planning / Study",
    "study": "Planning / Study",
    "planning / study": "Planning / Study",
    "waterres": "Water / Wastewater",
    "water resources": "Water / Wastewater",
    "waterww": "Water / Wastewater",
    "water/wastewater": "Water / Wastewater",
    "water / wastewater": "Water / Wastewater",
    "parks": "Park / Trail",
    "park": "Park / Trail",
    "trail": "Park / Trail",
    "park / trail": "Park / Trail",
    "traffic": "Traffic & Signals",
    "signals": "Traffic & Signals",
    "traffic & signals": "Traffic & Signals",
    "facilities": "Facilities",
    "buildings": "Facilities",
    "facilities & buildings": "Facilities",
    "facilities / buildings": "Facilities",
    "survey": "Survey & SUE",
    "sue": "Survey & SUE",
    "survey & sue": "Survey & SUE",
    "surveying": "Survey & SUE",
    "cei": "Construction Inspection",
    "inspection": "Construction Inspection",
    "construction inspection": "Construction Inspection",
    "technology": "Technology & GIS",
    "gis": "Technology & GIS",
    "technology & gis": "Technology & GIS",
    "row": "Right of Way",
    "right of way": "Right of Way",
    "drainage": "Drainage",
    "storm drainage": "Drainage",
    "drainage / stormwater": "Drainage",
    "environmental": "Environmental",
    "streambank stabilization": "Drainage",
    "intersection / traffic": "Traffic & Signals",
    "bridge / structural": "Bridge / Structural",
    "on-call": "On-Call",
    "other engineering": "Other Engineering",
}

# Category palette (category name → hex). Previously hardcoded in index.html;
# now emitted via config.json so the dashboard consumes a single source.
#
# Dict insertion order also drives the chart-display order (see config["order"]
# below — it's just `list(CATEGORY_PALETTE.keys())`). Drainage is intentionally
# placed above Facilities because drainage is a major Halff practice and
# should read as primary in the doughnut + heatmap + legend, not buried
# below smaller buckets.
#
# Color choices: kept the Halff brand tones for the four largest buckets
# (Blue / Teal / Salmon / Seafoam), then deliberately stepped away from the
# brand-blue cluster for the rest — a doughnut with 7 shades of blue is
# unreadable. Park = green (intuitive), Traffic = orange (intuitive),
# Facilities = plum, CEI = mustard, Survey = purple, Tech = cyan, ROW = rose,
# Env = brick, Bridge = slate, On-Call = warm gray. No two slices share a
# hex; the catch-all buckets (Other / Unknown) get the cool grays.
CATEGORY_PALETTE = {
    # Halff-toned palette. Halff Salmon (#FC6758) is intentionally NOT used as
    # a category color — it's reserved for highlighting Halff specifically
    # (Halff stats bar, Halff scatter point, Halff bar in Top 10, etc.).
    #
    # Top 4 are pure Halff brand colors. Roadway gets Maroon per Brent's
    # request — pops as the primary slice without using salmon. Lower-tier
    # categories use the supporting Halff-toned variants (mid seafoam, mauve,
    # light seafoam, amber, olive, purple, brick) so the whole palette
    # reads as Halff-branded.
    "Roadway":                 "#6F2740",  # Halff Maroon — primary, per Brent
    "Water / Wastewater":      "#115E6B",  # Halff Teal (water = blue family)
    "Drainage":                "#1C355E",  # Halff Blue (drainage = water = blue)
    "Planning / Study":        "#68949E",  # Halff Seafoam
    "Park / Trail":            "#4A7A8A",  # Mid seafoam (Halff-toned)
    "Traffic & Signals":       "#D19447",  # Amber
    "Facilities":              "#97536A",  # Mauve (Halff-toned)
    "Construction Inspection": "#8AAFB6",  # Light seafoam (Halff-toned)
    "Survey & SUE":            "#5E4B8B",  # Purple
    "Technology & GIS":        "#B7CECD",  # Halff Mint
    "Right of Way":            "#9B3426",  # Brick red
    "Environmental":           "#7A8B3F",  # Olive
    "Bridge / Structural":     "#3B5F7A",  # Navy variant (small bucket)
    "On-Call":                 "#6B7280",  # Medium cool gray
    "Other Engineering":       "#A0A0A0",  # Mid gray
    "Unknown":                 "#D9DAE4",  # Halff Cool Gray
}

# Halff canonical aliases — dashboard uses these for the "highlight Halff row"
# styling and top-firm rankings.
HALFF_ALIASES = ["halff associates", "halff"]


# Consultant name aliases — map alternate spellings, abbreviations, and
# acquired-firm pre-merger names to a single canonical form. Applied at
# JSON-export time in read_city() so the dashboard's bar charts, doughnut,
# and top-contracts table all reference one name per firm.
#
# Why each entry exists:
#   - "FNI": Freese and Nichols' internal abbreviation. Some council
#     documents use "FNI" in the action item even though the agreement
#     is awarded to "Freese and Nichols, Inc." — without this alias the
#     dashboard shows the firm under two names.
#   - "TRC Engineers, Inc." → Carollo Engineers, Inc.: TRC was acquired
#     by Carollo in March 2024. Contracts awarded after that point should
#     read Carollo; rows still entered under TRC are an artifact of how
#     the originating council document referenced the legacy name.
#     Folding TRC → Carollo is the post-acquisition reality.
#
# Add new aliases here as you encounter them. Keys are matched case-
# insensitively against the company text. Multiple alias forms can map
# to the same canonical name (e.g. "FNI" and "Freese & Nichols, Inc.").
CONSULTANT_ALIASES = {
    # Freese and Nichols
    "fni": "Freese and Nichols, Inc.",
    "freese & nichols, inc.": "Freese and Nichols, Inc.",
    "freese & nichols": "Freese and Nichols, Inc.",
    # Carollo (TRC acquisition, March 2024)
    "trc engineers, inc.": "Carollo Engineers, Inc.",
    "trc engineers": "Carollo Engineers, Inc.",
    "trc": "Carollo Engineers, Inc.",
}


def canonicalize_company(name: str) -> str:
    """Map FNI / TRC / etc. to their canonical post-merger names.

    Returns the original string when no alias applies.
    """
    if not name:
        return name
    key = name.strip().lower()
    return CONSULTANT_ALIASES.get(key, name)


# ── Row parsing helpers ─────────────────────────────────────────────────────


def normalize_type(raw):
    if not raw:
        return "Unknown"
    key = str(raw).strip().lower()
    if key in TYPE_MAP:
        return TYPE_MAP[key]
    return str(raw).strip().title()


# Drainage keyword override patterns. Used by reclassify_drainage() below to
# correct rows the upstream extractor (or human Excel entry) coded as
# "Planning / Study", "On-Call", or "Unknown" when the project name makes
# it obvious the work is drainage / floodplain / stormwater. Keep the
# pattern list specific — generic "stream" matches sweep in roadways.
_DRAINAGE_PROJECT_PATTERNS = (
    "flood mitigation",
    "flood study",
    "floodplain",
    "drainage and floodplain",
    "drainage study",
    "drainage master",
    "drainage improvement",
    "drainage engineering",
    "drainage design",
    "drainage relief",
    "drainage system",
    "stormwater",
    "storm water",
    "storm drain ",  # trailing space distinguishes from "storm drainage"
    "storm drainage",
    "detention pond",
    "stream restoration",
    "channel improvement",
    "channel design",
    "fema floodplain",
    "fema flood",
    "stream stabilization",
    "creek mitigation",
    "creek flood",
    # Hydraulic / hydrologic / H&H — these almost always indicate drainage
    # work in this dataset (flood-routing models, runoff calcs, etc.).
    "h&h study",
    "h & h study",
    "h&h analysis",
    "hydraulic",
    "hydrology",
    "hydrologic",
    "hydrologics",
    "hydraulics",
)

# Source categories that may be over-applied. We override these whenever the
# project NAME has an unambiguous drainage / floodplain / stormwater pattern.
# Water / Wastewater is included — Brent flagged the Halff/Dallas $5.99M Mill
# Creek storm drainage project that was sitting under W/WW. The fuzzy-match
# patterns above are conservative enough that water-replacement-with-incidental-
# storm-drain rows shouldn't trip the override (they don't lead with a drainage
# verb in the project title).
#
# Roadway and Park / Trail are intentionally NOT in the override list — a
# road project with drainage work or a park with stream restoration is still
# primarily that other discipline, and the project name usually reflects it.
_DRAINAGE_OVERRIDE_FROM = {
    "Planning / Study",
    "On-Call",
    "Unknown",
    "Other Engineering",
    "Water / Wastewater",
}


def reclassify_drainage(current_type: str, project: str, notes: str = "") -> str:
    """If a row's project NAME clearly identifies drainage / floodplain /
    stormwater work AND the current type is a generic catch-all (planning,
    on-call, unknown), return 'Drainage'. Otherwise return the current type
    unchanged.

    Only the project NAME is matched — notes often describe accessory work
    ("street reconstruction including drainage and utility upgrades") that
    would over-promote rows that are primarily roads or utilities.

    The override is conservative on purpose:
      - Won't promote Roadway → Drainage (a road project with a culvert is
        still primarily a road).
      - Won't promote Water / Wastewater → Drainage (cities legitimately
        bundle storm drain work into utility contracts).
      - Will promote Planning / Study → Drainage when a flood study or
        floodplain review was generic-coded as planning.
    """
    del notes  # intentionally unused — see docstring
    if current_type not in _DRAINAGE_OVERRIDE_FROM:
        return current_type
    name = (project or "").lower()
    for pat in _DRAINAGE_PROJECT_PATTERNS:
        if pat in name:
            return "Drainage"
    return current_type


def parse_date(val):
    """Convert Excel date value to ISO string."""
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.strftime("%Y-%m-%d")
    if isinstance(val, date):
        return val.strftime("%Y-%m-%d")
    s = str(val).strip()
    for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%m-%d-%Y"):
        try:
            return datetime.strptime(s, fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    return None


def read_city(slug: str, label: str, xlsx_path: str) -> list:
    """Read Config sheet from a city xlsx, return list of row dicts.

    Does NOT apply firm_filters — caller runs classify_contract() so the
    same policy applies uniformly and drop reasons are logged.
    """
    import openpyxl

    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    if "Config" not in wb.sheetnames:
        wb.close()
        raise RuntimeError("no Config sheet")

    ws = wb["Config"]
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    headers = [str(h).strip().lower() if h else "" for h in header_row]
    has_pm = "pm name" in headers
    pm_off = 1 if has_pm else 0

    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        year = row[0]
        date_val = parse_date(row[1])
        company = str(row[2]).strip() if row[2] else ""
        # Apply alias map so FNI/TRC/etc. roll up to canonical names.
        company = canonicalize_company(company)
        amount = row[3]
        project = str(row[4]).strip() if row[4] else ""
        proj_type = normalize_type(row[5])
        limits = str(row[6]).strip() if row[6] else ""
        notes = str(row[7]).strip() if row[7] else ""
        # Promote drainage / floodplain / stormwater / hydraulic-hydrologic
        # work that was generic-coded as planning / on-call / unknown.
        proj_type = reclassify_drainage(proj_type, project, notes)
        pm_name = str(row[8]).strip() if has_pm and len(row) > 8 and row[8] else ""
        src_file = (
            str(row[8 + pm_off]).strip()
            if len(row) > 8 + pm_off and row[8 + pm_off]
            else ""
        )
        pdf_idx = 10 + pm_off
        page_idx = 11 + pm_off
        pdf_link = (
            str(row[pdf_idx]).strip() if len(row) > pdf_idx and row[pdf_idx] else ""
        )
        page_num = row[page_idx] if len(row) > page_idx else None

        if not company and not amount:
            continue
        if year is not None and year < MIN_YEAR:
            continue
        try:
            amount = float(amount) if amount else 0
        except (ValueError, TypeError):
            amount = 0
        try:
            year = int(year) if year else None
        except (ValueError, TypeError):
            year = None

        rows.append(
            {
                "year": year,
                "date": date_val,
                "company": company,
                "amount": amount,
                "project": project,
                "type": proj_type,
                "limits": limits,
                "notes": notes,
                "pmName": pm_name,
                "srcFile": src_file,
                "city": label,
                "pdfLink": pdf_link,
                "pageNum": int(page_num) if page_num else None,
            }
        )

    wb.close()
    return rows


def apply_filters(rows: list) -> tuple[list, dict]:
    """Run each row through firm_filters.classify_contract().
    Returns (kept_rows, summary_counts).

    summary_counts = {"keep": N, "drop": M, "review": K, "drop_reasons": {...}}
    """
    kept = []
    counts = {"keep": 0, "drop": 0, "review": 0, "drop_reasons": {}}
    for r in rows:
        contract = {
            "company": r.get("company", ""),
            "amount": r.get("amount"),
            "project_name": r.get("project", ""),
            "project_type": r.get("type", ""),
            "description": r.get("notes", ""),
        }
        verdict, reason = classify_contract(contract)
        counts[verdict] = counts.get(verdict, 0) + 1
        if verdict == "keep":
            kept.append(r)
        elif verdict == "drop":
            # Collapse drop reasons to their leading phrase so summary is concise
            reason_key = reason.split(":", 1)[0].strip() if reason else "unknown"
            counts["drop_reasons"][reason_key] = (
                counts["drop_reasons"].get(reason_key, 0) + 1
            )
    return kept, counts


# Stop words for content-word signature (mirrors excel_writer's set so the
# JSON-level dedup behaves the same as the Excel-level pass).
_JSON_DEDUP_STOP_WORDS = frozenset(
    [
        "the", "and", "for", "with", "from", "into", "that", "this", "these",
        "those", "inc", "llc", "corp", "ltd", "co",
        "project", "projects", "services", "service", "agreement", "contract",
        "contracts", "engineering", "professional", "phase", "amendment",
        "design", "consulting", "consultants", "associates", "city", "award",
        "approve", "approval", "awarded", "approved", "authorize", "resolution",
        "ordinance", "work", "scope", "program", "plan", "planning", "study",
        "improvements", "improvement", "construction", "engineer", "engineers",
    ]
)


def _content_signature(text: str, n: int = 5) -> tuple:
    if not text:
        return ()
    tokens = re.findall(r"[a-z0-9]+", str(text).lower())
    content = [t for t in tokens if len(t) > 3 and t not in _JSON_DEDUP_STOP_WORDS]
    seen = list(dict.fromkeys(content))
    return tuple(sorted(seen[:n]))


def fuzzy_dedup_rows(rows: list, window_days: int = 90) -> tuple:
    """JSON-level fuzzy dedup. Catches cases where two rows share the same
    canonical company + amount bucket + project content words + city, within
    a ±window_days date band.

    Mirrors excel_writer._fuzzy_cross_sheet_dedup but operates on row dicts
    (post-canonicalize_company), so it catches FNI/Freese, TRC/Carollo, and
    other alias-driven duplicates that are dressed up under different firm
    names in the source Excel.

    Returns (kept_rows, removed_count).
    """
    if len(rows) < 2:
        return rows, 0

    # Build groups
    groups: dict = {}
    for idx, r in enumerate(rows):
        amt = r.get("amount") or 0
        if amt <= 0:
            continue  # leave $0 / on-call alone
        proj = r.get("project") or ""
        # Skip amendments — they're legitimate new contract actions that
        # happen to share firm + project text.
        proj_lower = proj.lower()
        if "amendment" in proj_lower or "change order" in proj_lower:
            continue
        date_str = r.get("date") or ""
        try:
            from datetime import datetime as _dt
            dt = _dt.strptime(date_str, "%Y-%m-%d") if date_str else None
        except ValueError:
            dt = None
        if dt is None:
            continue
        sig = (
            (r.get("company") or "").strip().lower(),
            int(round(amt / 1000.0)) * 1000,
            _content_signature(proj),
            (r.get("city") or "").strip().lower(),
        )
        groups.setdefault(sig, []).append({"idx": idx, "date": dt})

    # Collapse each group greedy: keep the latest, drop earlier rows within window
    drop_idxs = set()
    for sig, members in groups.items():
        if len(members) < 2:
            continue
        members.sort(key=lambda m: m["date"])
        # Walk backward, keeping the latest. Drop earlier rows within ±window_days.
        kept_dates = [members[-1]["date"]]
        for m in reversed(members[:-1]):
            nearest_kept = min(kept_dates, key=lambda d: abs((d - m["date"]).days))
            if abs((nearest_kept - m["date"]).days) <= window_days:
                drop_idxs.add(m["idx"])
            else:
                kept_dates.append(m["date"])

    if not drop_idxs:
        return rows, 0
    kept_rows = [r for i, r in enumerate(rows) if i not in drop_idxs]
    return kept_rows, len(drop_idxs)


def write_city_json(slug: str, label: str, rows: list, out_dir: str, xlsx_path: str) -> dict:
    """Write a single city's JSON file. Returns manifest entry.

    ``last_modified`` reflects the source Excel mtime (not the script-run
    time) so the dashboard can show a meaningful per-city "last processed"
    timestamp. Previously all cities shared the generator time, which made
    the per-city breakdown useless.
    """
    total_value = sum(r.get("amount") or 0 for r in rows)
    dates = [r["date"] for r in rows if r.get("date")]
    date_range = (
        {"min": min(dates), "max": max(dates)}
        if dates
        else {"min": None, "max": None}
    )
    payload = {
        "slug": slug,
        "name": label,
        "generated": datetime.now(timezone.utc).astimezone().isoformat(timespec="seconds"),
        "contract_count": len(rows),
        "total_value_usd": total_value,
        "contracts": rows,
    }
    path = os.path.join(out_dir, f"{slug}.json")
    with open(path, "w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False)
    size_bytes = os.path.getsize(path)
    try:
        xlsx_mtime_iso = datetime.fromtimestamp(
            os.path.getmtime(xlsx_path), tz=timezone.utc
        ).astimezone().isoformat(timespec="seconds")
    except OSError:
        # If we somehow can't stat the xlsx, fall back to the generator
        # time rather than omitting the field entirely.
        xlsx_mtime_iso = payload["generated"]
    return {
        "slug": slug,
        "name": label,
        "path": f"{slug}.json",
        "contract_count": len(rows),
        "total_value_usd": total_value,
        "date_range": date_range,
        "last_modified": xlsx_mtime_iso,
        "size_bytes": size_bytes,
        "status": "ok",
    }


def write_manifest(entries: list, out_path: str) -> None:
    """Write the manifest file. Entries include both ok + skipped cities."""
    manifest = {
        "generated": datetime.now(timezone.utc).astimezone().isoformat(timespec="seconds"),
        "generator": "build_dashboard_data.py v2.27",
        "cities": entries,
    }
    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(manifest, f, ensure_ascii=False, indent=2)


# ── PM aggregation (Item 26 phase 2) ──────────────────────────────────────

# Credential suffixes stripped when normalizing a PM name for dedup. Each is
# lowercase, no periods — applied after the raw name is lowercased + periods
# removed. "John A. Smith, P.E." → "john a smith pe" → strip "pe" →
# "john a smith". "J. Smith P.G." → "j smith pg" → "j smith".
_PM_CREDENTIALS = (
    "pe", "pmp", "aicp", "leed ap", "leed", "rla", "rlis",
    "ms", "msc", "mse", "phd", "pg", "se", "aia", "asla",
    "jr", "sr", "ii", "iii", "iv",
    "lpss", "rbpe", "gisp", "cfm",
)


def normalize_pm_key(name: str) -> str:
    """Collapse a raw PM name into a stable key for aggregation.

    Strips periods + parenthetical content + credential suffixes, lowercases,
    collapses whitespace. Intentionally does NOT collapse middle names or
    initials — "John A. Smith" and "John Smith" stay distinct until the user
    manually merges them. Conservative by design: merging on first+last alone
    would conflate real different people.
    """
    if not name:
        return ""
    s = str(name).strip().lower()
    # Remove parenthetical content — e.g. "John Smith (Halff)" → "john smith "
    s = re.sub(r"\([^)]*\)", "", s)
    # Drop periods so "P.E." matches "pe"
    s = s.replace(".", "")
    # Unify commas / hyphens / multiple whitespace → single space
    s = re.sub(r"[,\-\u2013\u2014]+", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    # Peel trailing credential tokens in a loop (someone might have multiple:
    # "John Smith, PE, PMP, LEED AP")
    while True:
        changed = False
        for cred in _PM_CREDENTIALS:
            if s.endswith(" " + cred):
                s = s[: -(len(cred) + 1)].strip()
                changed = True
                break
        if not changed:
            break
    return s


def _is_valid_pm_name(raw: str, key: str) -> bool:
    """Filter out extraction junk. Real PM names have 2+ words after normalize."""
    if not raw or not key:
        return False
    if len(raw) < 3 or len(raw) > 60:
        return False
    if len(key.split()) < 2:
        return False
    # Rejects file-path-y / obvious non-names
    bad_tokens = ("/", "\\", ".pdf", ".xlsx", "tbd", "n/a", "none", "unknown")
    low = raw.lower()
    if any(b in low for b in bad_tokens):
        return False
    return True


def aggregate_pms(all_kept_rows: list) -> list:
    """Group kept contracts by normalized PM key. Returns sorted PM list.

    Each entry carries:
      pm_key          — normalized key (internal dedup identifier)
      display_name    — most-frequent raw variant (what the UI shows)
      contract_count  — total rows attributed to this PM
      total_value_usd — sum of amounts
      avg_value_usd   — mean amount per contract
      firms           — [{name, count, value, date_range}], desc by value
                        (preserves firm-at-the-time so a PM who hopped
                        firms shows both tenures as separate rows)
      cities          — unique cities the PM has worked in
      types           — [{name, count}] sorted by frequency
    """
    pm_rows: dict = defaultdict(list)
    for c in all_kept_rows:
        raw = (c.get("pmName") or "").strip()
        key = normalize_pm_key(raw)
        if not _is_valid_pm_name(raw, key):
            continue
        pm_rows[key].append(c)

    out = []
    for key, rows in pm_rows.items():
        name_counts = Counter(r.get("pmName") for r in rows if r.get("pmName"))
        display_name = name_counts.most_common(1)[0][0] if name_counts else key.title()

        firm_groups: dict = defaultdict(list)
        for r in rows:
            firm_groups[r.get("company", "")].append(r)
        firms = []
        for firm_name, firm_rows in firm_groups.items():
            dates = [r["date"] for r in firm_rows if r.get("date")]
            firms.append(
                {
                    "name": firm_name,
                    "count": len(firm_rows),
                    "value": sum((r.get("amount") or 0) for r in firm_rows),
                    "date_range": {
                        "min": min(dates) if dates else None,
                        "max": max(dates) if dates else None,
                    },
                }
            )
        firms.sort(key=lambda f: -f["value"])

        cities = sorted({r.get("city", "") for r in rows if r.get("city")})
        type_counts = Counter(r.get("type", "") for r in rows if r.get("type"))
        types = [{"name": t, "count": n} for t, n in type_counts.most_common()]

        total_value = sum((r.get("amount") or 0) for r in rows)
        out.append(
            {
                "pm_key": key,
                "display_name": display_name,
                "contract_count": len(rows),
                "total_value_usd": total_value,
                "avg_value_usd": total_value / len(rows) if rows else 0,
                "firms": firms,
                "cities": cities,
                "types": types,
            }
        )

    out.sort(key=lambda p: (-p["total_value_usd"], -p["contract_count"]))
    return out


def write_pms_json(pms: list, coverage: dict, out_path: str) -> None:
    """Emit docs/data/pms.json for the PM Intelligence tab.

    ``coverage`` is a {"with_pm": N, "without_pm": M, "percent": X.X} dict
    so the dashboard can surface a "N of M contracts have PM data" badge.
    """
    payload = {
        "generated": datetime.now(timezone.utc).astimezone().isoformat(timespec="seconds"),
        "pm_count": len(pms),
        "total_pm_contracts": sum(p["contract_count"] for p in pms),
        "coverage": coverage,
        "pms": pms,
    }
    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False)


def write_dashboard_config(out_path: str) -> None:
    """Emit config.json from firm_filters + palette + halff aliases.

    The dashboard reads this to populate its defensive tripwire (arch firm
    detector) and the category palette, replacing hardcoded lists that
    previously drifted from firm_filters.py.
    """
    config = {
        "generated": datetime.now(timezone.utc).astimezone().isoformat(timespec="seconds"),
        "halff_aliases": HALFF_ALIASES,
        "categories": {
            "palette": CATEGORY_PALETTE,
            "order": list(CATEGORY_PALETTE.keys()),
        },
        # Defensive: dashboard logs a warning (not a filter) if any row
        # matches these after export. Normally nothing matches — we're
        # pre-filtering at export. Tripwire catches future regressions.
        "defensive_arch_firms": sorted(ARCHITECTURE_FIRMS),
        "defensive_non_engineering_firms": sorted(NON_ENGINEERING_FIRMS),
        "defensive_materials_geotech_firms": sorted(MATERIALS_GEOTECH_FIRMS),
    }
    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(config, f, ensure_ascii=False, indent=2)


def main():
    os.makedirs(OUTPUT_DIR, exist_ok=True)

    # Load city config
    with open(CONFIG_PATH, "r", encoding="utf-8") as f:
        config = json.load(f)
    cities = config.get("cities", {})

    manifest_entries = []
    all_kept_rows: list = []  # for cross-city PM aggregation
    total_kept = 0
    total_dropped = 0
    total_review = 0
    cross_city_drop_reasons: dict = {}

    for slug, info in cities.items():
        label = info.get("label", slug.title())
        if not info.get("enabled", True):
            manifest_entries.append(
                {
                    "slug": slug,
                    "name": label,
                    "path": None,
                    "status": "disabled",
                    "reason": "enabled=false in cities_config.json",
                }
            )
            continue

        xlsx_path = os.path.join(CONTRACT_LOGS_DIR, f"{slug}_contracts.xlsx")
        if not os.path.exists(xlsx_path):
            print(f"  SKIP {label:20s} file not found")
            manifest_entries.append(
                {
                    "slug": slug,
                    "name": label,
                    "path": None,
                    "status": "missing",
                    "reason": f"no file at {os.path.basename(xlsx_path)}",
                }
            )
            continue

        try:
            raw = read_city(slug, label, xlsx_path)
            kept, counts = apply_filters(raw)
            # Fuzzy dedup AFTER classify-and-filter has run, so we operate
            # on the canonicalized company names (FNI -> Freese and Nichols,
            # TRC -> Carollo). Catches duplicates that the per-Excel-sheet
            # fuzzy pass misses because the source rows are listed under
            # different firm names.
            kept, fuzzy_dropped = fuzzy_dedup_rows(kept)
            total_kept += counts.get("keep", 0) - fuzzy_dropped
            total_dropped += counts.get("drop", 0) + fuzzy_dropped
            total_review += counts.get("review", 0)
            for k, v in counts.get("drop_reasons", {}).items():
                cross_city_drop_reasons[k] = cross_city_drop_reasons.get(k, 0) + v
            if fuzzy_dropped:
                cross_city_drop_reasons["fuzzy_alias_dedup"] = (
                    cross_city_drop_reasons.get("fuzzy_alias_dedup", 0) + fuzzy_dropped
                )
            all_kept_rows.extend(kept)
            entry = write_city_json(slug, label, kept, OUTPUT_DIR, xlsx_path)
            manifest_entries.append(entry)
            summary = f"{entry['contract_count']:4d} contracts"
            if counts.get("drop"):
                summary += f"  ({counts['drop']} dropped)"
            if fuzzy_dropped:
                summary += f"  ({fuzzy_dropped} alias-deduped)"
            if counts.get("review"):
                summary += f"  ({counts['review']} to review)"
            print(f"  OK   {label:20s} {summary}")
        except PermissionError:
            print(f"  LOCK {label:20s} file open elsewhere")
            manifest_entries.append(
                {
                    "slug": slug,
                    "name": label,
                    "path": None,
                    "status": "locked",
                    "reason": "file open in another program",
                }
            )
        except Exception as e:
            print(f"  ERR  {label:20s} {e}")
            manifest_entries.append(
                {
                    "slug": slug,
                    "name": label,
                    "path": None,
                    "status": "error",
                    "reason": str(e)[:200],
                }
            )

    write_manifest(manifest_entries, MANIFEST_PATH)
    write_dashboard_config(DASHBOARD_CONFIG_PATH)

    # ── PM aggregation for the PM Intelligence tab ──
    pms = aggregate_pms(all_kept_rows)
    with_pm = sum(p["contract_count"] for p in pms)
    without_pm = len(all_kept_rows) - with_pm
    pct = (with_pm / len(all_kept_rows) * 100) if all_kept_rows else 0.0
    coverage = {
        "with_pm": with_pm,
        "without_pm": without_pm,
        "percent": round(pct, 1),
    }
    write_pms_json(pms, coverage, PMS_PATH)
    print(
        f"PMs aggregated: {len(pms)} unique PMs, "
        f"{with_pm}/{len(all_kept_rows)} contracts have PM "
        f"({pct:.1f}% coverage)"
    )

    ok_cities = [e for e in manifest_entries if e.get("status") == "ok"]
    skipped = [e for e in manifest_entries if e.get("status") != "ok"]

    print()
    print(f"Output:       {OUTPUT_DIR}/")
    print(f"Cities ok:    {len(ok_cities)} ({', '.join(e['name'] for e in ok_cities)})")
    if skipped:
        print(
            "Cities skip:  "
            + f"{len(skipped)} "
            + ", ".join(f"{e['name']} [{e['status']}]" for e in skipped)
        )
    print(f"Total kept:   {total_kept}")
    print(f"Total dropped:{total_dropped}")
    print(f"Total review: {total_review}")
    if cross_city_drop_reasons:
        print("Drop reasons:")
        for reason, n in sorted(
            cross_city_drop_reasons.items(), key=lambda x: -x[1]
        ):
            print(f"  {n:4d}  {reason}")


if __name__ == "__main__":
    main()
