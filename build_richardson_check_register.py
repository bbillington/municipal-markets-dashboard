"""
Build docs/richardson_check_register.json from the Check Register xlsx.

Richardson sources contract data through two different paths:
  1. Council agenda packets (handled by build_dashboard_data.py) — almost
     always empty because Richardson's Sec. 2-52(d) delegates most engineering
     contract awards to City Manager administrative action (no Council
     approval required, no public agenda entry).
  2. AP Check Register (this script) — the actual record of what got paid.
     Sourced from discovery.cor.gov, aggregated in
     Contract Logs/richardson_check_register.xlsx.

The two data paths have DIFFERENT schemas and must NOT be merged. This script
produces a standalone JSON that the dashboard consumes via a Richardson-only
section.
"""

import json
import os
import sys
from datetime import date, datetime

if sys.platform == "win32":
    sys.stdout.reconfigure(encoding="utf-8")

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
REPO_ROOT = os.path.dirname(SCRIPT_DIR)
XLSX_PATH = os.path.join(REPO_ROOT, "Contract Logs", "richardson_check_register.xlsx")
OUTPUT_PATH = os.path.join(SCRIPT_DIR, "richardson_check_register.json")

# Apply the same outlier-toggleable list from build_dashboard_data.py.
# These architecture firms and PM vendors are flagged so the dashboard
# can toggle them on/off alongside the Council-packet outliers.
OUTLIER_TOGGLEABLE_FIRMS = [
    "perkins & will",
    "dlr group",
    "inspire dallas",
]
OUTLIER_TOGGLEABLE_THRESHOLDS = [
    {"match": "gensler", "min_amount": 5_000_000},
    {"match": "kai/alliance", "min_amount": 5_000_000},
    {"match": "kai design", "min_amount": 5_000_000},
]


def _is_toggleable_outlier(company: str, amount: float = 0) -> bool:
    if not company:
        return False
    name = company.lower()
    for firm in OUTLIER_TOGGLEABLE_FIRMS:
        if firm in name:
            return True
    try:
        amt = float(amount) if amount else 0
    except (ValueError, TypeError):
        amt = 0
    for rule in OUTLIER_TOGGLEABLE_THRESHOLDS:
        if rule["match"] in name and amt >= rule["min_amount"]:
            return True
    return False


def _to_iso(val):
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.strftime("%Y-%m-%d")
    if isinstance(val, date):
        return val.strftime("%Y-%m-%d")
    return str(val)


def main():
    import openpyxl

    if not os.path.exists(XLSX_PATH):
        print(f"ERROR: Check Register not found at {XLSX_PATH}")
        sys.exit(1)

    wb = openpyxl.load_workbook(XLSX_PATH, read_only=True, data_only=True)

    # Summary sheet: Vendor, # Payments, Total Paid, First Payment, Last Payment, Departments
    vendors = []
    ws = wb["Summary"]
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True)):
        if not row or not row[0]:
            continue
        vendor, n_payments, total, first, last, departments = (row + (None,) * 6)[:6]
        try:
            n_payments = int(n_payments) if n_payments is not None else 0
        except (ValueError, TypeError):
            n_payments = 0
        try:
            total = float(total) if total is not None else 0.0
        except (ValueError, TypeError):
            total = 0.0
        v = {
            "vendor": str(vendor).strip(),
            "payments": n_payments,
            "total": total,
            "first": _to_iso(first),
            "last": _to_iso(last),
            "departments": str(departments).strip() if departments else "",
        }
        if _is_toggleable_outlier(v["vendor"], v["total"]):
            v["outlier"] = True
        vendors.append(v)

    # Details sheet: vendor x fiscal month
    details = []
    ws = wb["Details"]
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        vendor, fm, n_payments, total, departments = (row + (None,) * 5)[:5]
        try:
            n_payments = int(n_payments) if n_payments is not None else 0
        except (ValueError, TypeError):
            n_payments = 0
        try:
            total = float(total) if total is not None else 0.0
        except (ValueError, TypeError):
            total = 0.0
        d = {
            "vendor": str(vendor).strip(),
            "fiscalMonth": str(fm).strip() if fm else "",
            "payments": n_payments,
            "total": total,
            "departments": str(departments).strip() if departments else "",
        }
        if _is_toggleable_outlier(d["vendor"], d["total"]):
            d["outlier"] = True
        details.append(d)

    wb.close()

    # Sort vendors by total desc
    vendors.sort(key=lambda v: v["total"], reverse=True)

    total_paid = sum(v["total"] for v in vendors)
    total_payments = sum(v["payments"] for v in vendors)

    output = {
        "generated": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "city": "Richardson",
        "source": "AP Check Register (discovery.cor.gov)",
        "provenance_note": (
            "Richardson procurement follows Sec. 2-52(d) — the City Manager "
            "has administrative authority to approve engineering PSAs without "
            "Council action. Council agenda packets therefore contain almost "
            "no engineering contract awards. This tab is sourced from the AP "
            "Check Register and shows actual payments (not awards), so totals "
            "are cumulative across multi-year contracts."
        ),
        "vendor_count": len(vendors),
        "total_payments": total_payments,
        "total_paid": total_paid,
        "vendors": vendors,
        "details": details,
    }

    with open(OUTPUT_PATH, "w", encoding="utf-8") as f:
        json.dump(output, f, ensure_ascii=False)

    print(f"Output: {OUTPUT_PATH}")
    print(f"Vendors:  {len(vendors)}")
    print(f"Details:  {len(details)} vendor-month rows")
    print(f"Payments: {total_payments:,}")
    print(f"Paid:     ${total_paid:,.0f}")
    print(f"Outlier-flagged vendors: {sum(1 for v in vendors if v.get('outlier'))}")
    print(f"File size: {os.path.getsize(OUTPUT_PATH) / 1024:.0f} KB")


if __name__ == "__main__":
    main()
